# services/employee_service.py

from database import get_company_collection, get_employee_collection
from bson.objectid import ObjectId
from fastapi import HTTPException
import pandas as pd
import math
import json
from utils.excel_utils import validate_excel_columns, create_excel_from_employees_with_formulas
from utils.query_utils import build_query_filters
from models.employee import EmployeeCreate, EmployeeUpdate
from typing import Dict, Any, List, Optional
from io import BytesIO

class JSONEncoder(json.JSONEncoder):
    """Custom JSON encoder to handle NaN, Infinity, and -Infinity values."""
    def default(self, obj):
        if isinstance(obj, float):
            if math.isnan(obj):
                return None
            if math.isinf(obj):
                return None
        return super().default(obj)

def clean_nan_values(data):
    """
    Recursively replace NaN, Infinity, and -Infinity values with None in dictionaries and lists.
    
    Args:
        data: The data structure to clean (dict, list, or primitive value)
        
    Returns:
        The cleaned data structure with NaN values replaced by None
    """
    if isinstance(data, dict):
        return {k: clean_nan_values(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [clean_nan_values(item) for item in data]
    elif isinstance(data, float) and (math.isnan(data) or math.isinf(data)):
        return None
    else:
        return data

async def upload_employees(company_id: str, file):
    from utils.excel_extraction import extract_columns_from_excel
    from utils.excel_utils import validate_excel_columns
    from io import BytesIO

    def sanitize_field_name(name: str) -> str:
        return (
            name.replace(".", "")
                .replace("$", "")
                .replace("\n", " ")
                .strip()
                .lower()
                .replace("  ", " ")
                .replace(" ", "_")
        )

    company = await get_company_collection().find_one({"_id": ObjectId(company_id)})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    expected_columns = company["salary_sheet_columns"]

    try:
        excel_bytes = await file.read()
        file.file.seek(0)
        excel_file = BytesIO(excel_bytes)

        # Extract header info
        header_info = await extract_columns_from_excel(file)
        columns = header_info["columns"]
        header_row = header_info["header_row_index"]

        file.file.seek(0)
        df = pd.read_excel(excel_file, header=header_row)

        if not validate_excel_columns(columns, expected_columns):
            raise HTTPException(
                status_code=400,
                detail=f"Column mismatch: Expected {expected_columns}, got {columns}"
            )

        # Clean rows
        df.dropna(how='all', inplace=True)
        df = df[~df.apply(lambda row: row.astype(str).str.lower().str.contains("total").any(), axis=1)]
        df = df.replace({pd.NA: None, float('nan'): None, float('inf'): None, float('-inf'): None})

        employee_collection = get_employee_collection()
        inserted, updated, skipped = 0, 0, 0

        # Find email column
        email_col = next((c for c in columns if "email" in c.lower()), None)
        if not email_col:
            raise HTTPException(status_code=400, detail="No 'email' column found in uploaded sheet.")

        sanitized_email_col = sanitize_field_name(email_col)

        # Columns to exclude from employee collection
        attendance_columns = [
            "no. of days in a month",
            "no. of days present"
        ]
        attendance_columns_normalized = [sanitize_field_name(col) for col in attendance_columns]

        for _, row in df.iterrows():
            sanitized_employee = {}
            for col in columns:
                norm_col = sanitize_field_name(col)
                if norm_col in attendance_columns_normalized:
                    continue  # Skip attendance columns
                if col in row:
                    sanitized_employee[norm_col] = row[col]
            sanitized_employee["company_id"] = company_id
            ...

            sanitized_employee["company_id"] = company_id

            email_value = sanitized_employee.get(sanitized_email_col, "")
            if not email_value:
                skipped += 1
                continue

            email_value = str(email_value).strip().lower()

            existing = await employee_collection.find_one({
                "company_id": company_id,
                sanitized_email_col: {"$regex": f"^{email_value}$", "$options": "i"}
            })

            if existing:
                await employee_collection.update_one(
                    {"_id": existing["_id"]},
                    {"$set": sanitized_employee}
                )
                updated += 1
            else:
                await employee_collection.insert_one(sanitized_employee)
                inserted += 1

        return {
            "message": "Employee upload completed",
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")



async def add_employee(employee_data: EmployeeCreate):
    collection = get_company_collection()
    company = await collection.find_one({"_id": ObjectId(employee_data.company_id)})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Validate that the employee data contains all the required columns
    for column in company["salary_sheet_columns"]:
        if column not in employee_data.data and column.lower() not in [k.lower() for k in employee_data.data.keys()]:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required field: {column}"
            )
    
    # Create the employee document
    employee_doc = employee_data.data
    employee_doc["company_id"] = employee_data.company_id
    
    # Clean NaN values
    employee_doc = clean_nan_values(employee_doc)
    
    try:
        result = await get_employee_collection().insert_one(employee_doc)
        return {
            "message": "Employee added successfully",
            "employee_id": str(result.inserted_id)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

async def get_employees(company_id: str, skip: int = 0, limit: int = 100, filters: Optional[Dict[str, Any]] = None):
    # Validate company exists
    company = await get_company_collection().find_one({"_id": ObjectId(company_id)})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Build query
    query = {"company_id": company_id}
    if filters:
        # Convert filters to MongoDB query
        mongo_filters = build_query_filters(filters)
        query.update(mongo_filters)
    
    # Execute query
    cursor = get_employee_collection().find(query).skip(skip).limit(limit)
    employees = await cursor.to_list(length=limit)
    
    # Get total count for pagination
    total_count = await get_employee_collection().count_documents(query)
    
    # Convert ObjectId to string and clean NaN values
    for employee in employees:
        employee["_id"] = str(employee["_id"])
    
    # Clean NaN values to make JSON serializable
    employees = clean_nan_values(employees)
    
    return {"employees": employees, "total": total_count}

async def get_employee(company_id: str, employee_id: str):
    try:
        employee = await get_employee_collection().find_one({
            "_id": ObjectId(employee_id),
            "company_id": company_id
        })
        
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        employee["_id"] = str(employee["_id"])
        
        # Clean NaN values
        employee = clean_nan_values(employee)
        
        return employee
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

async def update_employee(company_id: str, employee_id: str, employee_data: EmployeeUpdate):
    # Check if employee exists
    employee = await get_employee_collection().find_one({
        "_id": ObjectId(employee_id),
        "company_id": company_id
    })
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Clean NaN values
    cleaned_data = clean_nan_values(employee_data.data)
    
    try:
        await get_employee_collection().update_one(
            {"_id": ObjectId(employee_id)},
            {"$set": cleaned_data}
        )
        return {"message": "Employee updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

async def delete_employee(company_id: str, employee_id: str):
    # Check if employee exists
    employee = await get_employee_collection().find_one({
        "_id": ObjectId(employee_id),
        "company_id": company_id
    })
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    try:
        await get_employee_collection().delete_one({"_id": ObjectId(employee_id)})
        return {"message": "Employee deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

async def export_employees(company_id: str, filters: Optional[Dict[str, Any]] = None):
    """
    Export employees to Excel with properly maintained formulas.
    
    Args:
        company_id: ID of the company
        filters: Optional filters to apply to the query
        
    Returns:
        BytesIO stream containing the Excel file
    """
    # Validate company exists
    company = await get_company_collection().find_one({"_id": ObjectId(company_id)})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    # Build query
    query = {"company_id": company_id}
    if filters:
        # Convert filters to MongoDB query
        mongo_filters = build_query_filters(filters)
        query.update(mongo_filters)
    
    # Get columns from company
    columns = company.get("salary_sheet_columns", [])
    
    # Get formula mapping if available
    formula_mapping = company.get("salary_sheet_formulas", {})
    
    # If no formulas are available, try to infer them based on the column names
    if not formula_mapping:
        # Define default formulas for common calculated fields
        default_formulas = {}
        
        # Get column indices for reference
        column_indices = {col.lower(): idx for idx, col in enumerate(columns)}
        
        # Attempt to identify key columns
        basic_pay_idx = next((idx for col, idx in column_indices.items() if "basic pay" in col), None)
        hra_idx = next((idx for col, idx in column_indices.items() if "hra" in col), None)
        
        # Gross Amount = sum of all compensation components
        gross_amount_idx = next((idx for col, idx in column_indices.items() 
                             if "gross" in col and "amount" in col), None)
        
        # Total Deductions = sum of all deduction components
        total_ded_idx = next((idx for col, idx in column_indices.items() 
                          if "total" in col and "ded" in col), None)
        
        # Net Amount = Gross Amount - Total Deductions
        net_amt_idx = next((idx for col, idx in column_indices.items() 
                         if "net" in col and "amt" in col), None)
        
        if gross_amount_idx is not None:
            # Find all allowance columns
            allowance_cols = []
            for col, idx in column_indices.items():
                if any(term in col for term in ["basic", "hra", "allow", "education", "reimb", "lta", 
                                               "medical", "attire", "sp.all"]):
                    # Make sure this isn't the gross amount itself
                    if idx != gross_amount_idx:
                        col_name = columns[idx]
                        col_letter = get_column_letter(idx + 1)  # +1 because Excel is 1-indexed
                        allowance_cols.append(f"{col_letter}{{row}}")
            
            if allowance_cols:
                gross_col_name = columns[gross_amount_idx]
                default_formulas[gross_col_name] = "=" + "+".join(allowance_cols)
        
        if total_ded_idx is not None:
            # Find all deduction columns
            deduction_cols = []
            for col, idx in column_indices.items():
                if any(term in col for term in ["tax", "tds", "p. f", "pf", "esic", "advance", "deduction"]):
                    # Make sure this isn't the total deduction itself
                    if idx != total_ded_idx:
                        col_name = columns[idx]
                        col_letter = get_column_letter(idx + 1)
                        deduction_cols.append(f"{col_letter}{{row}}")
            
            if deduction_cols:
                total_ded_col_name = columns[total_ded_idx]
                default_formulas[total_ded_col_name] = "=" + "+".join(deduction_cols)
        
        if gross_amount_idx is not None and total_ded_idx is not None and net_amt_idx is not None:
            gross_col_letter = get_column_letter(gross_amount_idx + 1)
            ded_col_letter = get_column_letter(total_ded_idx + 1)
            net_amt_col_name = columns[net_amt_idx]
            
            default_formulas[net_amt_col_name] = f"={gross_col_letter}{{row}}-{ded_col_letter}{{row}}"
        
        formula_mapping = default_formulas
    
    # Fetch employees
    cursor = get_employee_collection().find(query)
    employees = await cursor.to_list(length=None)  # Get all matching employees
    
    if not employees:
        raise HTTPException(status_code=404, detail="No employees found")
    
    # Clean NaN values
    employees = clean_nan_values(employees)
    
    # Import the column letter function
    from openpyxl.utils import get_column_letter
    
    # Generate Excel with the company's format and formulas
    try:
        file_stream = create_excel_from_employees_with_formulas(
            employees, 
            columns, 
            company["name"],
            formula_mapping
        )
        return file_stream
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Error generating Excel: {str(e)}"
        )
