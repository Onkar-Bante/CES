# routers/company_router.py
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from models.company import CompanyCreate, CompanyUpdateSalaryFormat
from services.company_service import create_company, update_salary_format, extract_and_update_salary_format
from utils.excel_utils import generate_sample_template
from fastapi.responses import StreamingResponse
from io import BytesIO
from bson.objectid import ObjectId
from database import get_company_collection
from typing import List

router = APIRouter()

@router.post("/create_company")
async def api_create_company(company: CompanyCreate):
    return await create_company(company)

@router.put("/update_salary_format/{company_id}")
async def api_update_salary_format(company_id: str, data: CompanyUpdateSalaryFormat):
    return await update_salary_format(company_id, data)

@router.post("/extract_salary_columns/{company_id}")
async def api_extract_salary_columns(company_id: str, file: UploadFile = File(...)):
    """
    Extract salary sheet columns from an uploaded Excel template
    and update the company's salary format
    """
    return await extract_and_update_salary_format(company_id, file)

@router.post("/upload_salary_template/{company_id}")
async def import_salary_sheet_with_formulas(company_id: str, file: UploadFile = File(...)):
    """
    Import a salary sheet Excel file with formulas and update the company's salary format.
    
    Args:
        company_id: ID of the company
        file: Uploaded Excel template file
        
    Returns:
        Dict with success message
    """
    try:
        # Read the file
        contents = await file.read()
        file.file.seek(0)  # Reset file pointer
        
        # First, extract the columns
        from utils.excel_extraction import extract_columns_from_excel
        header_info = await extract_columns_from_excel(file)
        columns = header_info["columns"]
        header_row_idx = header_info["header_row_index"]
        
        file.file.seek(0)  # Reset file pointer
        
        # Use openpyxl to read the workbook with formulas
        from openpyxl import load_workbook
        
        excel_file = BytesIO(contents)
        workbook = load_workbook(excel_file, data_only=False)
        sheet = workbook.active
        
        # Extract formulas from data rows
        formulas = {}
        data_row_start = header_row_idx + 1
        
        # Check multiple rows to ensure we capture all formulas
        for data_row_idx in range(data_row_start + 1, min(data_row_start + 5, sheet.max_row + 1)):
            for col_idx in range(1, sheet.max_column + 1):
                header_cell = sheet.cell(row=data_row_start, column=col_idx)
                if header_cell.value:
                    header = str(header_cell.value).strip()
                    
                    cell = sheet.cell(row=data_row_idx, column=col_idx)
                    if cell.data_type == 'f' and cell.value and str(cell.value).startswith('='):
                        formula = str(cell.value)
                        # Replace the specific row number with {row} template
                        formula = formula.replace(str(data_row_idx), "{row}")
                        
                        # Don't overwrite if we already found a formula for this column
                        if header not in formulas:
                            formulas[header] = formula
        
        # Update company document
        await get_company_collection().update_one(
            {"_id": ObjectId(company_id)},
            {"$set": {
                "salary_sheet_columns": columns,
                "salary_sheet_formulas": formulas
            }}
        )
        
        return {
            "message": "Successfully imported salary sheet format and formulas", 
            "columns_count": len(columns),
            "formulas_count": len(formulas),
            "columns": columns,
            "formulas": formulas
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error importing salary sheet with formulas: {str(e)}")

@router.get("/generate_template/{company_id}")
async def generate_template(company_id: str):
    from database import get_company_collection
    from bson.objectid import ObjectId

    company = await get_company_collection().find_one({"_id": ObjectId(company_id)})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    columns = company["salary_sheet_columns"]
    file_stream = generate_sample_template(columns)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={company_id}_template.xlsx"}
    )