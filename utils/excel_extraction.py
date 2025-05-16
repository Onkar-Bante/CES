# utils/excel_extraction.py
import pandas as pd
from io import BytesIO
from fastapi import UploadFile, HTTPException

async def extract_columns_from_excel(file: UploadFile, header_row_index: int = None) -> dict:
    try:
        contents = await file.read()
        file.file.seek(0)
        excel_file = BytesIO(contents)

        indicators = ["sr", "emp", "id", "name", "email", "basic", "hra", "gross", "net", "tax", "deduction"]
        
        if header_row_index is not None:
            df = pd.read_excel(excel_file, header=header_row_index)
            cleaned_columns = clean_columns(df.columns)
            return {"columns": cleaned_columns, "header_row_index": header_row_index}

        preview_df = pd.read_excel(excel_file, header=None, nrows=10)
        candidates = []

        for i in range(len(preview_df)):
            row = preview_df.iloc[i]
            row_strs = row.fillna("").astype(str).str.lower()
            keyword_matches = sum(1 for cell in row_strs if any(ind in cell for ind in indicators))
            non_empty_cells = sum(1 for cell in row_strs if cell.strip() and "unnamed" not in cell)

            if keyword_matches >= 3 and non_empty_cells >= len(row) * 0.5:
                candidates.append((i, keyword_matches))

        if candidates:
            best_row = max(candidates, key=lambda x: x[1])[0]
            file.file.seek(0)
            df = pd.read_excel(file.file, header=best_row)
            cleaned_columns = clean_columns(df.columns)
            return {"columns": cleaned_columns, "header_row_index": best_row}

        # Fallback to row 3
        file.file.seek(0)
        df = pd.read_excel(file.file, header=2)
        cleaned_columns = clean_columns(df.columns)
        return {"columns": cleaned_columns, "header_row_index": 2}

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to extract headers: {str(e)}")

def clean_columns(columns):
    cleaned = []
    for i, col in enumerate(columns):
        if pd.isna(col) or "unnamed" in str(col).lower():
            cleaned.append(f"Column_{i+1}")
        else:
            cleaned.append(str(col).strip())  # Already strips leading/trailing spaces
    return cleaned

    
async def extract_formulas_from_excel(file: UploadFile) -> dict:
    """
    Extract Excel formulas from an uploaded file with improved detection.
    
    Args:
        file: Uploaded Excel file
        
    Returns:
        Dictionary mapping column names to formula templates
    """
    try:
        # Read the file
        contents = await file.read()
        file.file.seek(0)  # Reset file pointer
        
        # First detect the header row
        header_info = await extract_columns_from_excel(file)
        header_row_index = header_info["header_row_index"]
        columns = header_info["columns"]
        
        file.file.seek(0)  # Reset file pointer again
        
        # Use openpyxl to read the workbook with formulas
        from openpyxl import load_workbook
        
        excel_file = BytesIO(contents)
        workbook = load_workbook(excel_file, data_only=False)
        sheet = workbook.active
        
        # Get headers using the detected header row
        headers = {}
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=header_row_index + 1, column=col_idx)
            if cell.value:
                headers[col_idx] = str(cell.value).strip()
        
        # Look for formulas in the first few data rows (try multiple rows to ensure we find formulas)
        formulas = {}
        
        # Check up to 5 rows after the header row to find formulas
        for data_row in range(header_row_index + 2, min(header_row_index + 7, sheet.max_row + 1)):
            for col_idx, header in headers.items():
                cell = sheet.cell(row=data_row, column=col_idx)
                if cell.data_type == 'f' and cell.value and str(cell.value).startswith('='):
                    formula = str(cell.value)
                    # Replace the specific row number with {row} template
                    formula = formula.replace(str(data_row), "{row}")
                    
                    # Don't overwrite if we already found a formula for this column
                    if header not in formulas:
                        formulas[header] = formula
        
        # If we didn't find any formulas, try to infer common ones based on column names
        if not formulas:
            # First identify potential formula columns (e.g., "gross amount", "total ded", "net amt")
            calculated_columns = [
                col for col in columns 
                if any(term in col.lower() for term in [
                    "gross", "total", "net", "payable", "deduction", "subtotal", "sum"
                ])
            ]
            
            # For each potential formula column, create a basic formula
            for col in calculated_columns:
                col_lower = col.lower()
                
                # Find the column index
                col_idx = None
                for idx, header in headers.items():
                    if header.lower() == col_lower:
                        col_idx = idx
                        break
                
                if not col_idx:
                    continue
                
                # Get the column letter
                col_letter = sheet.cell(row=1, column=col_idx).column_letter
                
                # Create formulas based on column name patterns
                if "gross" in col_lower and "amount" in col_lower:
                    # For "Gross Amount": sum of basic pay + HRA + allowances
                    input_cols = [c for c in columns if any(
                        term in c.lower() for term in [
                            "basic", "hra", "allow", "reimb", "conv", "lta", "medical", "education"
                        ]
                    )]
                    
                    if input_cols:
                        # Find column letters for input columns
                        input_letters = []
                        for input_col in input_cols:
                            for idx, header in headers.items():
                                if header.lower() == input_col.lower():
                                    letter = sheet.cell(row=1, column=idx).column_letter
                                    input_letters.append(f"{letter}{{row}}")
                                    break
                        
                        if input_letters:
                            formulas[col] = "=" + "+".join(input_letters)
                
                elif "total" in col_lower and "ded" in col_lower:
                    # For "Total Ded": sum of tax + deductions
                    deduction_cols = [c for c in columns if any(
                        term in c.lower() for term in [
                            "tax", "tds", "esic", "p.f", "pf", "advance", "deduction"
                        ]
                    )]
                    
                    if deduction_cols:
                        # Find column letters for deduction columns
                        deduction_letters = []
                        for deduct_col in deduction_cols:
                            for idx, header in headers.items():
                                if header.lower() == deduct_col.lower():
                                    letter = sheet.cell(row=1, column=idx).column_letter
                                    deduction_letters.append(f"{letter}{{row}}")
                                    break
                        
                        if deduction_letters:
                            formulas[col] = "=" + "+".join(deduction_letters)
                
                elif "net" in col_lower and "amt" in col_lower:
                    # For "Net Amt": Gross Amount - Total Deductions
                    gross_col = next((c for c in columns if "gross" in c.lower() and "amount" in c.lower()), None)
                    total_ded_col = next((c for c in columns if "total" in c.lower() and "ded" in c.lower()), None)
                    
                    if gross_col and total_ded_col:
                        gross_letter = None
                        ded_letter = None
                        
                        for idx, header in headers.items():
                            if header.lower() == gross_col.lower():
                                gross_letter = sheet.cell(row=1, column=idx).column_letter
                            elif header.lower() == total_ded_col.lower():
                                ded_letter = sheet.cell(row=1, column=idx).column_letter
                        
                        if gross_letter and ded_letter:
                            formulas[col] = f"={gross_letter}{{row}}-{ded_letter}{{row}}"
                
                elif "payable" in col_lower:
                    # For "Payable": Net Amt + Other Reimbursements + Bonus
                    net_col = next((c for c in columns if "net" in c.lower() and "amt" in c.lower()), None)
                    bonus_cols = [c for c in columns if any(
                        term in c.lower() for term in ["bonus", "reimbursement", "other"]
                    )]
                    
                    if net_col:
                        net_letter = None
                        for idx, header in headers.items():
                            if header.lower() == net_col.lower():
                                net_letter = sheet.cell(row=1, column=idx).column_letter
                                break
                        
                        formula_parts = [f"{net_letter}{{row}}"]
                        
                        for bonus_col in bonus_cols:
                            for idx, header in headers.items():
                                if header.lower() == bonus_col.lower():
                                    letter = sheet.cell(row=1, column=idx).column_letter
                                    formula_parts.append(f"{letter}{{row}}")
                                    break
                        
                        if formula_parts:
                            formulas[col] = "=" + "+".join(formula_parts)
        
        return formulas
        
    except Exception as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=f"Error extracting formulas: {str(e)}")