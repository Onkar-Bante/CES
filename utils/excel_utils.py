# utils/excel_utils.py

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import math

def validate_excel_columns(actual_columns, expected_columns):
    """Validate that all expected columns are present in the actual columns."""
    # Normalize column names for comparison (lowercase, strip whitespace)
    normalized_actual = [col.lower().strip() for col in actual_columns]
    normalized_expected = [col.lower().strip() for col in expected_columns]
    
    # Check if all expected columns are present
    return all(exp in normalized_actual for exp in normalized_expected)

def generate_excel_template(columns):
    """Generate an empty Excel template with the specified columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Sheet Template"
    
    # Add title row
    ws.merge_cells('A1:F1')
    ws['A1'] = "Salary Sheet Template"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Add instructions row
    ws.merge_cells('A2:F2')
    ws['A2'] = "Fill in employee data below. All columns are required."
    ws['A2'].font = Font(italic=True)
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    header_row = 3
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Set column width based on column name length
        ws.column_dimensions[chr(64 + col_idx)].width = max(len(column) + 5, 15)
    
    # Add sample row formatting
    sample_row = header_row + 1
    for col_idx in range(1, len(columns) + 1):
        ws.cell(row=sample_row, column=col_idx)
    
    # Save to stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_sample_template(columns, company_name):
    """Generate a sample Excel template with example data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Salary Sheet"
    
    # Add title row
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{company_name} - Sample Salary Template"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Add instructions row
    ws.merge_cells('A2:F2')
    ws['A2'] = "This is a sample file with example data. Replace with your real employee data."
    ws['A2'].font = Font(italic=True)
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    header_row = 3
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Set column width based on column name length
        ws.column_dimensions[chr(64 + col_idx)].width = max(len(column) + 5, 15)
    
    # Add sample data rows
    sample_data = [
        {
            "EMP ID": "EMP001",
            "Name of Employees": "John Doe",
            "Email": "john.doe@example.com",
            "Designation": "Manager",
            "Name of Site": "Main Branch",
            "Basic Pay": 50000,
            "HRA": 15000,
            "DA": 5000,
            "Special Allowance": 8000,
            "Gross Amt": 78000,
            "PF": 6000,
            "TDS": 5500,
            "Net Amt": 66500
        },
        {
            "EMP ID": "EMP002",
            "Name of Employees": "Jane Smith",
            "Email": "jane.smith@example.com",
            "Designation": "Developer",
            "Name of Site": "Tech Center",
            "Basic Pay": 45000,
            "HRA": 13500,
            "DA": 4500,
            "Special Allowance": 7000,
            "Gross Amt": 70000,
            "PF": 5400,
            "TDS": 4900,
            "Net Amt": 59700
        }
    ]
    
    # Fill in sample data
    for idx, data in enumerate(sample_data):
        row_idx = header_row + idx + 1
        for col_idx, column in enumerate(columns, start=1):
            value = data.get(column, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def create_excel_from_employees_with_formulas(employees, columns, company_name, formula_mapping=None):
    """
    Create an Excel file from employees data that includes formulas for calculated fields.
    
    Args:
        employees: List of employee documents
        columns: List of column names
        company_name: Name of the company
        formula_mapping: Dictionary mapping column names to Excel formula templates
        
    Returns:
        BytesIO containing the Excel file
    """
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from datetime import datetime
    import calendar
    import re
    
    # Create a DataFrame with only the specified columns in the correct order
    data = []
    
    # Identify which columns need formulas
    formula_columns = set()
    if formula_mapping:
        formula_columns = {col.lower() for col in formula_mapping.keys()}
    
    # Specified input-only columns based on your example
    # This is inferred from your comment about "this columns only have data rest are empty"
    input_only_columns = {
        "name of employees", "email", "designation", "name of site", 
        "no. of days in a month", "no. of days present", "basic pay", 
        "hra", "conv allow/ vehicle reimb", "child education allow", 
        "lta monthly", "medical monthly", "attire reimb", "sp.all", "other allow"
    }
    
    def normalize_key(key: str) -> str:
        import re
        key = re.sub(r'\s+', ' ', key)  # Replace all whitespace with single space
        key = key.replace('\n', ' ').replace('\r', ' ')
        key = key.strip().lower()
        key = re.sub(r'[\s\./\\-]+', '_', key)  # Replace space, dot, slash, backslash, dash with underscore
        key = re.sub(r'_+', '_', key)           # Replace multiple underscores with one
        key = re.sub(r'[^a-z0-9_]', '', key)    # Remove all non-alphanumeric except underscore
        return key

    for emp in employees:
        # Normalize employee keys for lookup
        normalized_emp = {normalize_key(k): v for k, v in emp.items()}
        row = {}
        for col in columns:
            col_norm = normalize_key(col)
            if col_norm in formula_columns:
                row[col] = None
            else:
                row[col] = normalized_emp.get(col_norm, None)
        data.append(row)
        # print(normalized_emp.keys())
        # print([normalize_key(col) for col in columns])
    
    df = pd.DataFrame(data, columns=columns)
    
    # Create Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Add title rows
        current_date = datetime.now()
        month_year = f"{calendar.month_name[current_date.month]}- {str(current_date.year)[2:]}"
        
        # Write to excel starting from row 3 to leave space for the header
        df.to_excel(writer, index=False, startrow=2)
        
        # Get the worksheet
        worksheet = writer.sheets['Sheet1']
        
        # Add company name
        worksheet.merge_cells('A1:D1')
        cell = worksheet.cell(row=1, column=1)
        cell.value = f"{company_name} SALARY SHEET"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='left')
        
        # Add month-year
        worksheet.merge_cells('E1:G1')
        cell = worksheet.cell(row=1, column=5)
        cell.value = month_year
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='left')
        
        # Style header row (row 3)
        for idx, col in enumerate(columns, 1):
            cell = worksheet.cell(row=3, column=idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Add a border
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            cell.border = thin_border
            
            # Auto-adjust column width based on content
            max_length = len(col)
            for row_idx in range(4, len(df) + 4):  # +4 because data starts at row 4
                cell_value = worksheet.cell(row=row_idx, column=idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = max(15, max_length + 2)
        
        # Apply formulas
        if formula_mapping:
            for row_idx in range(4, len(df) + 4):  # +4 because data starts at row 4
                for col_name, formula_template in formula_mapping.items():
                    # Find the column index for this formula
                    col_idx = None
                    for i, col in enumerate(columns, 1):
                        if col.lower() == col_name.lower():
                            col_idx = i
                            break
                    
                    if col_idx:
                        # Replace {row} in formula with current row
                        formula = formula_template.replace("{row}", str(row_idx))
                        
                        # Apply the formula to the cell
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.value = formula
                        cell.data_type = 'f'  # Set as formula
                        
        # Add total row at the bottom
        total_row_idx = len(df) + 4
        worksheet.cell(row=total_row_idx, column=1).value = "TOTAL"
        worksheet.cell(row=total_row_idx, column=1).font = Font(bold=True)
        
        # Add SUM formulas for numerical columns
        for idx, col in enumerate(columns, 1):
            col_lower = col.lower()
            if any(term in col_lower for term in ["amount", "pay", "amt", "basic", "hra", "conv", "reimb", "allow"]):
                col_letter = worksheet.cell(row=3, column=idx).column_letter
                cell = worksheet.cell(row=total_row_idx, column=idx)
                start_cell = f"{col_letter}4"
                end_cell = f"{col_letter}{total_row_idx-1}"
                cell.value = f"=SUM({start_cell}:{end_cell})"
                cell.font = Font(bold=True)
                cell.number_format = "#,##0.00"
    
    output.seek(0)
    return output

def apply_formulas_to_row(worksheet, row_idx, formula_mapping, columns):
    """Apply Excel formulas to a specific row."""
    for col_idx, column in enumerate(columns, start=1):
        if column in formula_mapping:
            formula = formula_mapping[column]
            # Replace {row} template with actual row number
            formula = formula.replace("{row}", str(row_idx))
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = formula
            
def add_table_styling(worksheet, header_row, rows_count, cols_count):
    """Add borders and styling to make the data look like a table."""
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Apply borders to all data cells
    for row in range(header_row, header_row + rows_count + 1):
        for col in range(1, cols_count + 1):
            worksheet.cell(row=row, column=col).border = thin_border
            
            # Center-align numbers for better readability
            cell = worksheet.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")